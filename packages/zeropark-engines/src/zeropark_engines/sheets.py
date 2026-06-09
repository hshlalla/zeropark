"""SHEETS — render table data to a real .xlsx. Native (openpyxl).

Content generation (turning a prompt into table data) is an LLM step;
OpenpyxlSheetsEngine renders whatever table data it is given.
LLMSheetsEngine orchestrates the LLM step and rendering.
"""

from __future__ import annotations

import json
import re
import asyncio
import io
from typing import Any
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from zeropark_core.capabilities import Capability
from zeropark_core import ArtifactStore
from zeropark_core.models import Artifact, TaskRequest, TaskResult, TaskStatus
from zeropark_core.llm import BaseLLMClient, ChatMessage

from zeropark_engines.base import NativeEngine

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class OpenpyxlSheetsEngine(NativeEngine):
    id = "openpyxl-sheets"
    name = "Openpyxl Renderer Engine"
    capabilities = frozenset({Capability.SHEETS})
    reference = "Genspark AI Sheets - Excel Renderer"

    def __init__(self, store: ArtifactStore) -> None:
        self.store = store

    def _render_xlsx(self, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        if headers:
            ws.append(headers)
            for i, _ in enumerate(headers, 1):
                ws[f"{get_column_letter(i)}1"].font = Font(bold=True)

        for row in rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def cap_sheets(self, task: TaskRequest, task_id: str) -> TaskResult:
        # Expected structure: {"headers": ["col1", "col2"], "rows": [[val1, val2], ...]}
        sheet_data = task.params.get("sheet_data") or {
            "headers": ["A"],
            "rows": [[task.prompt[:80]]]
        }
        deck_title = task.params.get("title") or task.prompt[:80]

        xlsx_bytes = self._render_xlsx(
            deck_title, 
            sheet_data.get("headers", []), 
            sheet_data.get("rows", [])
        )
        
        filename = f"{task_id}.xlsx"
        file_uri = self.store.save(filename, xlsx_bytes)

        n_rows = len(sheet_data.get("rows", []))
        artifact = Artifact(
            id=f"{task_id}_sheet",
            kind="sheet",
            title=deck_title,
            mime_type=_XLSX_MIME,
            uri=file_uri,
        )
        return TaskResult(
            task_id=task_id,
            status=TaskStatus.SUCCEEDED,
            capability=Capability.SHEETS,
            provider_id=self.id,
            artifacts=[artifact],
            metrics={"rows": n_rows},
        )


class LLMSheetsEngine(NativeEngine):
    id = "llm-sheets"
    name = "LLM Sheets Generator"
    capabilities = frozenset({Capability.SHEETS})
    reference = "Genspark AI Sheets - LLM Table Generation"

    def __init__(self, llm_client: BaseLLMClient, renderer: OpenpyxlSheetsEngine, model_name: str = "gpt-4o-mini") -> None:
        self.llm_client = llm_client
        self.renderer = renderer
        self.model_name = model_name

    async def _generate_sheet_data(self, prompt: str) -> dict:
        system_msg = ChatMessage(
            role="system",
            content=(
                "You are an expert spreadsheet designer. Create a table of data based on the user's prompt. "
                "The data must be in a strict JSON format with exactly two keys: 'headers' (a list of strings) and 'rows' (a list of lists containing strings, numbers, or excel formulas like '=SUM(B2:B5)'). "
                "Structure example: {\"headers\": [\"Item\", \"Cost\"], \"rows\": [[\"A\", 10], [\"B\", 20], [\"Total\", \"=SUM(B2:B3)\"]]}. "
                "Return ONLY the JSON object, nothing else."
            )
        )
        user_msg = ChatMessage(role="user", content=prompt)
        
        loop = asyncio.get_event_loop()
        response = await loop.run_in_executor(
            None,
            lambda: self.llm_client.chat_completion([system_msg, user_msg], self.model_name, temperature=0.3)
        )
        
        text = response.content.strip()
        # Find JSON object using regex if there's markdown wrapping
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            text = match.group(0)
            
        try:
            sheet_data = json.loads(text)
            if isinstance(sheet_data, dict) and "headers" in sheet_data and "rows" in sheet_data:
                return sheet_data
        except Exception:
            pass
            
        # Fallback if parsing fails
        return {"headers": ["Error"], "rows": [["Could not generate structured table."]]}

    async def cap_sheets(self, task: TaskRequest, task_id: str) -> TaskResult:
        # 1. Generate Table Data via LLM
        sheet_data = await self._generate_sheet_data(task.prompt)
        
        # 2. Inject data into task params
        if not task.params:
            task.params = {}
        task.params["sheet_data"] = sheet_data
        
        # 3. Delegate to renderer
        return await self.renderer.cap_sheets(task, task_id)
